import copy

from tqdm import tqdm, trange

import torch
from torch import optim
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
import torch.nn as nn
import numpy as np
import pathlib
import pandas as pd
import openpyxl

from diffusion_model import DiffusionProcess
from utils import *


from models.unet import *

def run():
    batch = 44
    train_paths = get_paths_to_patient_files_val('C:/Dataset/train')
    val_paths = get_paths_to_patient_files_val('C:/Dataset/valid')

    #Dataload code add
    #
    #
    #
    #
    #


    # Prepare model and training
    device = "cuda"
    gan_model = UNet_Gan().to(device)
    ddpm_model = GeneratorUNet(in_channels=7,n_cls=7,n_filters=32).to(device)

    process = DiffusionProcess(device=device)
    optimizerG = optim.Adam(gan_model.parameters(), lr=1e-3)
    schedulerG = optim.lr_scheduler.StepLR(optimizerG, 80)

    optimizer = optim.Adam(ddpm_model.parameters(), lr=1e-3)
    scheduler = optim.lr_scheduler.StepLR(optimizer, 80)

    criterion = torch.nn.MSELoss()
    criterionG = nn.MSELoss()

    learning_curves = dict()
    learning_curves['loss'] = dict()
    learning_curves['loss']['train'], learning_curves['loss']['val'] = [], []

    best_val_epoch = 0
    best_val_loss = float('inf')
    best_val_avg_loss = float('inf')
    best_model_wts = None
    checkpoint = None

    os.makedirs("./Results", exist_ok=True)
    path_to_save_dir = './Results/'

    excel = openpyxl.Workbook()
    excel_path = path_to_save_dir + 'loss_and_metrics.xlsx'
    excel.save(excel_path)


    epoch = 200
    for e in trange(epoch):
        for phase in ['train', 'val']:
            if phase == 'train':


                with torch.set_grad_enabled(phase == 'train'):
                    phase_noise_mse = 0.0
                    phase_img_nmse = 0.0
                    phase_total_loss = 0.0
                    train_bar = tqdm(trainloader,leave=False)
                    for data in train_bar:
                        id = data['id']
                        ac_img, nac_img = data['ac_img'].to(device), data['nac_img'].to(device)

                        gan_model.train()
                        optimizerG.zero_grad()

                        fake_ac_img = gan_model(nac_img)
                        img_loss = criterionG(fake_ac_img,ac_img)

                        img_loss.backward()
                        optimizerG.step()

                        ddpm_model.train()
                        optimizer.zero_grad()

                        fake_ac_img = gan_model(nac_img)
                        t = torch.randint(0, 100, (ac_img.shape[0],)).to(device)
                        epsilon = torch.randn(fake_ac_img.shape).to(device)
                        diffused_image, comb_noise = process.forward(ac_img, fake_ac_img, t, epsilon)

                        predict_noise = ddpm_model(diffused_image.to(device), t.to(device))
                        noise_loss = criterion(comb_noise.to(device), predict_noise)
                        phase_noise_mse += noise_loss.item()
                        phase_img_nmse += img_loss.item()

                        total_loss = img_loss+noise_loss

                        noise_loss.backward()
                        optimizer.step()

                        train_bar.set_description(f"Noise_loss: {noise_loss.item():.4f} Img_loss: {img_loss.item():.4f} total_loss: {total_loss.item():.4f}")
                        del noise_loss, img_loss, total_loss

                    schedulerG.step()
                    scheduler.step()
                    phase_noise_mse /= len(trainloader)
                    phase_img_nmse /= len(trainloader)
                    phase_total_loss /= len(trainloader)

                    excel_ = openpyxl.load_workbook(excel_path)
                    excel_ws = excel_.active
                    excel_ws.cell(row=e + 1, column=1).value = phase_noise_mse
                    excel_ws.cell(row=e + 1, column=2).value = phase_img_nmse
                    excel_ws.cell(row=e + 1, column=3).value = phase_total_loss
                    excel_.save(excel_path)

                    tqdm.write(f"Mean loss for Epoch {e +1}: {phase_total_loss:.4f} {phase_noise_mse:.4f} {phase_img_nmse:.4f}")
                    learning_curves['loss']['train'].append(phase_total_loss)


            else:
                gan_model.eval()
                ddpm_model.eval()

                phase_img_nmse = 0.0
                phase_noise_mse = 0.0
                phase_total_loss = 0.0

                with torch.set_grad_enabled(phase == 'train'):
                    val_bar = tqdm(valloader,leave=False)
                    for data in val_bar:
                        id = data['id']
                        ac_img, nac_img = data['ac_img'], data['nac_img']
                        ac_img, nac_img = ac_img.to(device), nac_img.to(device)

                        xt_ = torch.randn(batch, 7, 200, 200).to(device)
                        f_ac =gan_model(nac_img)

                        for t in trange(99,-1,-1):
                            time = torch.ones(batch) * t
                            et_ = ddpm_model(xt_.to(device), time.to(device))
                            xt_ = process.proposed_inverse(xt_,f_ac, et_, t)

                        loss_mse =torch.nn.L1Loss()(xt_,ac_img)
                        phase_img_nmse += loss_mse.item()


                    phase_img_nmse /= len(valloader)

                    excel_ = openpyxl.load_workbook(excel_path)
                    excel_ws = excel_.active
                    excel_ws.cell(row=e + 1, column=3).value = phase_img_nmse
                    excel_.save(excel_path)

                    tqdm.write(f"Mean img loss for Epcoh {e + 1}: {phase_img_nmse:.4f}")
                    learning_curves['loss']['val'].append(phase_img_nmse)

                    if phase_img_nmse < best_val_loss:
                        best_val_epoch = e
                        best_val_loss = phase_img_nmse
                        best_model_wts = copy.deepcopy(gan_model.state_dict())
                        best_model_wts2 = copy.deepcopy(ddpm_model.state_dict())

        if (e + 1) % 10 == 0:
           torch.save(best_model_wts, os.path.join('Results/', str(e) + '_' + str(best_val_epoch) + '_gan_best_model_weights.pt'))
           torch.save(best_model_wts2,
                      os.path.join('Results/', str(e) + '_' + str(best_val_epoch) + '_ddpm_best_model_weights.pt'))

        if (e + 1) % 1 == 0:
            torch.save(gan_model.state_dict(), os.path.join('Results/', str(e) + '_gan_model_weights.pt'))
            torch.save(ddpm_model.state_dict(), os.path.join('Results/', str(e) + '_ddpm_model_weights.pt'))

    path_to_dir = pathlib.Path('Results')

    # Check if the directory exists:
    if not os.path.exists(path_to_dir):
        os.makedirs(path_to_dir)

    # Write a short summary in a csv file:
    with open(path_to_dir / 'summary.csv', 'w', newline='', encoding='utf-8') as summary:
        summary.write(f'SUMMARY OF THE EXPERIMENT:\n\n')
        summary.write(f'BEST VAL EPOCH: {best_val_epoch}\n')
        summary.write(f'BEST VAL LOSS: {best_val_loss}\n')

    # Save best model weights:
    torch.save(best_model_wts, path_to_dir / 'best_model_weights.pt')

    df_learning_curves = pd.DataFrame.from_dict({
        'loss_train': learning_curves['loss']['train'],
        'loss_val': learning_curves['loss']['val']
    })
    df_learning_curves.to_csv(path_to_dir / 'learning_curves.csv', sep=';')

    # Save learning curves' plots in png files:
    # Loss figure:
    plt.figure(figsize=(17.5, 10))
    plt.plot(range(epoch), learning_curves['loss']['train'], label='train')
    plt.plot(range(epoch), learning_curves['loss']['val'], label='val')
    plt.xlabel('Epoch', fontsize=20)
    plt.ylabel('Loss', fontsize=20)
    plt.xticks(fontsize=15)
    plt.yticks(fontsize=15)
    plt.legend(fontsize=20)
    plt.grid()
    plt.savefig(path_to_dir / 'loss_plot.png', bbox_inches='tight')


if __name__=="__main__":
    run()

